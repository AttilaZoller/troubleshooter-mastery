from bs4 import BeautifulSoup
import re
from fuzzywuzzy import process
import openpyxl


STEAM_MASTERY_REGEX = r"(Mastery Set - |AC\.MS - | Class | Type |\(.*\))"


class troubleshooterMastery:
    def __init__(self):
        self.korean_dictionary, self.english_dictionary = self._collect_dictionary()
        corrected_dictionary_workbook = openpyxl.load_workbook(
            "corrected_code_dictionary.xlsx", data_only=True
        )
        corrected_dictionary_worksheet = corrected_dictionary_workbook["Sheet1"]
        self.raw_text_dictionary = [
            cell.value for cell in list(corrected_dictionary_worksheet["A:A"])
        ]
        self.corrected_text_dictionary = [
            cell.value for cell in list(corrected_dictionary_worksheet["B:B"])
        ]
        self.code_dictionary = [
            cell.value for cell in list(corrected_dictionary_worksheet["C:C"])
        ]

    def main(self):
        with open("guide.html", encoding="utf-8", mode="r") as f:
            html = f.read()

        soup = BeautifulSoup(html, "html.parser")
        section_contents = soup.select("div.subSection.detailBox")

        category = {
            "mastery_category_name": "",
            "mastery_set_name": "",
            "mastery_set_components": "",
        }

        try:
            workbook = openpyxl.load_workbook("Mastery.xlsx", data_only=True)
            worksheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(["카테고리 분류", "세트명", "특성 1", "특성 2", "특성 3", "특성 4"])

        for section in section_contents:
            sub_section_title = str(
                section.select("div.subSectionTitle")[0].text
            ).strip()

            if "Mastery Set" in sub_section_title or "AC.MS" in sub_section_title:
                category["mastery_category_name"] = re.sub(
                    STEAM_MASTERY_REGEX, "", sub_section_title
                )

                mastery_set_name_contents = section.select("div.subSectionDesc > div")
                mastery_contents = section.select("ul")
                for index, mastery in enumerate(mastery_contents):
                    try:
                        category["mastery_set_name"] = (
                            mastery_set_name_contents[index]
                            .text.replace("-", "")
                            .strip()
                        )
                        category["mastery_set_components"] = (
                            mastery.select("li > b")[0].text.strip().split(" + ")
                        )
                    except IndexError:
                        category["mastery_set_name"] = (
                            mastery_set_name_contents[index]
                            .text.strip()
                            .replace("- ", "")
                        )
                        category["mastery_set_components"] = (
                            section.select("div.subSectionDesc > b")[0]
                            .text.strip()
                            .split(" + ")
                        )
                    finally:
                        category["mastery_set_components"] = [
                            mastery.strip()
                            for mastery in category["mastery_set_components"]
                        ]
                        english_data = [
                            category["mastery_category_name"],
                            category["mastery_set_name"],
                        ] + category["mastery_set_components"]
                        korean_data = self.translate_english_to_korean(english_data)
                        worksheet.append(korean_data)
                        workbook.save("Mastery.xlsx")

    def _collect_dictionary(self):
        with open("dic_keyword.dic", encoding="utf-8", mode="r") as f:
            raw_text = f.readlines()

        korean_dictionary = []
        english_dictionary = []
        for line in raw_text:
            if "!Mastery" in line or "!Job" in line or "!AbilitySubType" in line:
                tag_name = re.compile(r"(?<=\!).*?(?=\])").search(line).group()

                mastery_tags = re.compile(r"{}]".format(tag_name)).findall(line)
                if len(mastery_tags) == 2:
                    korean_name = (
                        re.compile(r"(?<={}]).*?(?=\[\!{}])".format(tag_name, tag_name))
                        .search(line)
                        .group()
                        .strip()
                    )
                    parsed_text = line.replace(
                        "{}]{}".format(tag_name, korean_name), ""
                    )
                    english_name = (
                        re.compile(r"(?<={}]).*".format(tag_name))
                        .search(parsed_text)
                        .group()
                        .strip()
                    )

                elif len(mastery_tags) == 1:
                    korean_name = (
                        re.compile(r"(?<={}]).*?(?=[a-zA-Z])".format(tag_name))
                        .search(line)
                        .group()
                        .strip()
                    )
                    parsed_text = line.replace("{}".format(korean_name), "")
                    english_name = (
                        re.compile(r"(?<={}]).*".format(tag_name))
                        .search(parsed_text)
                        .group()
                        .strip()
                    )

                korean_dictionary.append(korean_name.replace(" ", " "))
                english_dictionary.append(english_name.replace(" ", " "))

        return korean_dictionary, english_dictionary

    def translate_english_to_korean(self, english_data):
        korean_data = []

        for data in english_data:
            data = data.replace(" ", " ")
            if (data in self.raw_text_dictionary) == True:
                index = self.raw_text_dictionary.index(data)
                data = self.corrected_text_dictionary[index]

            highest_match = process.extractOne(data, self.english_dictionary)
            translated_word = self.korean_dictionary[
                self.english_dictionary.index(highest_match[0])
            ]

            if highest_match[1] == 100:
                korean_data.append(translated_word)

            elif highest_match[1] > 90:
                with open("translating_log.txt", mode="a", encoding="utf-8") as f:
                    f.write("{}\n".format(data))
                korean_data.append(translated_word)

            elif highest_match[1] <= 90:
                with open("translating_log.txt", mode="a", encoding="utf-8") as f:
                    f.write("{}\n".format(data))
                    # f.write("{}는 {}({})와(과) 일치율 {}(으)로 영어명을 그대로 사용함.\n".format(data,highest_match[0],translated_word,str(highest_match[1])))
                korean_data.append(data)

        return korean_data


mastery = troubleshooterMastery()
mastery.main()
